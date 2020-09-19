#Importing all modules
import cv2, time, datetime, numpy, pandas, os, statistics, imutils, shutil, sys, json
import openpyxl as xl
from xml.dom import minidom
from datetime import timedelta
from imutils.video import FileVideoStream
from bokeh.plotting import figure, show, output_file
from bokeh.models import HoverTool, ColumnDataSource
from PyQt5.QtWidgets import QMessageBox, QFileDialog
from PyQt5.QtGui import QPixmap, QImage
from PyQt5.QtCore import QSize
    
def check_pos(first, second, thresh):
    f_x = first[0]
    f_y = first[1]
    s_x = second[0]
    s_y = second[1]
    if f_x - s_x <= thresh and f_x - s_x >= -thresh:
        if f_y - s_y <= thresh and f_y - s_y >= -thresh:
            return True
        else:
            return False
    else:
        return False

def analyse(videopath, xmlpath, folderpath, VideoID_List, Window):
    """Analyse a video. videopath(String): Full path to the processed video. xmlpath(String): Full path to the XML File. folderpath(String): Full path to the folder to store results in. VideoID_List(String): ID of the processing video. Window(Qt Window object): The Window where the UI for the analysation is in."""
    Window.analysation_status_image.setMovie(Window.loading_animation)
    Window.loading_animation.start()

    #Reading Files
    video = FileVideoStream(videopath).start()
    black = cv2.imread('files/black.png', 0)

    try:
        os.rename(src=f'{folderpath}/frames', dst=f'{folderpath}/remove')
    except:
        pass

    shutil.rmtree(f'{folderpath}/remove', ignore_errors=True)

    try:
        os.mkdir(f'{folderpath}/frames')
    except FileExistsError:
        pass
    
    #Defining the variable for the reference frame
    ref_frame = None

    start_meteor = [[0,0]]
    global position_names
    position_names = []
    global MeteorID_List
    MeteorID_List = []
    rotation_list = []
    Meteor_area_list = []
    
    status_list=[None,None]
    times=[]
    
    #Reading and processing the XML-file
    XML_File = minidom.parse(xmlpath)
    creationDate = XML_File.getElementsByTagName('CreationDate')
    creationDate = creationDate[0].attributes['value'].value

    length = XML_File.getElementsByTagName('Duration')
    length = length[0].attributes['value'].value
    length = int(length)

    Fps = XML_File.getElementsByTagName('VideoFrame')
    Fps = Fps[0].attributes['captureFps'].value
    Fps = int(Fps[:-1])

    Width = XML_File.getElementsByTagName('VideoLayout')
    Width = Width[0].attributes['pixel'].value
    Width = int(Width)

    Height = XML_File.getElementsByTagName('VideoLayout')
    Height = Height[0].attributes['numOfVerticalLine'].value
    Height = int(Height)

    year = int(creationDate[:4])
    month = int(creationDate[5:7])
    day = int(creationDate[8:10])
    hour = int(creationDate[11:13])
    minute = int(creationDate[14:16])
    second = int(creationDate[17:19])

    base_time = datetime.datetime(year, month, day, hour, minute, second)
    beginning_video_time = datetime.datetime(1, 1, 1, 0, 0, 0)
    
    len_mul = Height//1080
    ar_mul = len_mul**2

    #Defining the variable for the current frame
    frame_number = 1
    
    #Processing the Excel-File
    global wb
    wb = xl.load_workbook('files/Results_template.xlsx')
    sheet = wb['Daten']
    base = "0000000"
    meteor_count = 0
    detection_count = 0

    meteor_data = {}
    
    for i in range(length):
        progress_percent = frame_number / length * 100
        Window.analysation_progressbar.setProperty("value", progress_percent)

        #Reading current frame
        frame = video.read()

        #Reset variables
        status=0

        #Preparing the Video for calculations
        gray = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)

        gray = cv2.GaussianBlur(gray,(20*len_mul+1, 20*len_mul+1),0)

        if ref_frame is None:
            ref_frame = gray
            print("Reference frame is first frame.")
            status_list=[0]
            continue
        if frame_number%1000 == 0:
            ref_frame = gray
            print("Frame reset after 1000 frames")

        #Launch the draw_grid function
        x_grid = 8
        y_grid = 5

        width = Width//x_grid
        height = Height//y_grid
        for v in range(y_grid):
            move = v*height
            for h in range(x_grid):
                cv2.rectangle(frame, (width*h,move), (width*(h+1),move+height), (100,100,100), 1)

        #Calculating the difference between the current frame and the reference frame
        delta_frame = cv2.absdiff(ref_frame, gray)

        thresh_delta = cv2.threshold(delta_frame, 20, 255, cv2.THRESH_BINARY)[1]

        thresh_delta = cv2.dilate(thresh_delta, None, iterations=2)

        (cnts,_) = cv2.findContours(thresh_delta.copy(),cv2.RETR_EXTERNAL,cv2.CHAIN_APPROX_SIMPLE)

        #Calculating the time
        current_seconds = frame_number/Fps

        time = base_time + timedelta(seconds=current_seconds)

        #Drawing boxes around areas that are meteors
        if len(cnts) > 5:
            ref_frame = gray
            print("Frame was reset because more than 5 Meteors")
        else:
            for contour in cnts:
                if cv2.contourArea(contour) < 60*ar_mul or cv2.contourArea(contour) > 4000*ar_mul:
                    continue
                status=1
                detection_count += 1
                Meteor_area_list.append(cv2.contourArea(contour))
                rect = cv2.minAreaRect(contour)
                rotation_list.append(rect[2])
                (x, y, w, h) = cv2.boundingRect(contour)
                border = 10*len_mul
                text_border = 20*len_mul
                cv2.rectangle(frame, (x-border,y-border), (x+w+border,y+h+border),(255,255,255), 2)
                cv2.putText(frame, "Meteor", (x-border,y-text_border), cv2.FONT_HERSHEY_SIMPLEX, 1*len_mul,(100,255,0),1,cv2.LINE_AA)
                average_x = x+(w//2)
                average_y = y+(h//2)
                meteor_data[f"{frame_number}_meteor{detection_count}"] = {"position" : (average_x, average_y), "frame" : [frame_number]}


        #Return if a meteor was detected
        status_list.append(status)
        status_list=status_list[-3:]

        #Return the Start and End time of the meteor
        if status_list[-1]==1 and status_list[-2]==0: #If the meteor appeared
            meteor_count+=1
            count_digits = len(str(meteor_count))
            times.append(time)

            start_meteor= [average_x, average_y]

            width_box = Width//x_grid
            height_box = Height//y_grid

            X_names = ["A", "B", "C", "D", "E", "F", "G", "H", "I", "J", "L", "M", "N", "O", "P"]

            for i in range(x_grid):
                if start_meteor[0] > (i+1)*width_box:
                    continue
                X_position_name = X_names[i]
                break

            for i in range(y_grid):
                if start_meteor[1] > (i+1)*height_box:
                    continue
                Y_position_name = str(i+1)
                break

            Position_name = X_position_name + Y_position_name

            position_names.append(Position_name)
            position_names.append(Position_name)

            current_meteor_start = base_time + timedelta(seconds = current_seconds)

            row_number = meteor_count+1
            cell = sheet.cell(row_number,1)
            MeteorID = "M-"+base[:-count_digits-1]+str(meteor_count)
            MeteorID_List.append(MeteorID)
            MeteorID_List.append(MeteorID)
            cell.value = MeteorID

            cell = sheet.cell(row_number,2)
            cell.value = str(base_time)[-8:]

            cell = sheet.cell(row_number,3)
            cell.value = VideoID_List

            cell = sheet.cell(row_number,4)
            cell.value = str(beginning_video_time + timedelta(seconds = current_seconds))[11:]

            cell = sheet.cell(row_number,6)
            cell.value = str(current_meteor_start)[11:]

            cell = sheet.cell(row_number,11)
            cell.value = Position_name

            cell = sheet.cell(row_number,12)
            cell.value = str(current_meteor_start)[:10]



        if status_list[-1]==0 and status_list[-2]==1: #If the meteor disappeared
            current_meteor_end = base_time + timedelta(seconds = current_seconds)
            meteor_length = current_meteor_end - current_meteor_start

            if meteor_length > timedelta(seconds=5) or meteor_length < timedelta(seconds=0.08):
                meteor_count += -1
                times.pop() #Remove the beginning from the list
                position_names.pop()
                position_names.pop()
                sheet.delete_rows(row_number)
            else:
                times.append(time)

                cell = sheet.cell(row_number,5)
                cell.value = str(beginning_video_time + timedelta(seconds = current_seconds))[11:]

                cell = sheet.cell(row_number,7)
                cell.value = str(current_meteor_end)[11:]

                cell = sheet.cell(row_number,8)
                cell.value = str(current_meteor_end - current_meteor_start)

                cell = sheet.cell(row_number,9)
                Meteor_area_list = sorted(Meteor_area_list)
                Meteor_area = Meteor_area_list[-1]
                Meteor_area = round(Meteor_area)
                cell.value = Meteor_area

                cell = sheet.cell(row_number,10)
                rot_average = statistics.mean(rotation_list)
                cell.value = str(round(rot_average))+"°"

                rotation_list = []
                Meteor_area_list = []

        #Resize the main window to match the screen size
        frameResized = cv2.resize(frame, (1280, 720))

        #Writing the text for the time, date and frame number
        cv2.putText(frameResized, str(time)[:-5], (20, 700), cv2.FONT_HERSHEY_DUPLEX, 0.5,(200,200,200),1,cv2.LINE_AA)
        cv2.putText(frameResized, str(frame_number), (1220, 700), cv2.FONT_HERSHEY_DUPLEX, 0.5,(200,200,200),1,cv2.LINE_AA)

        #Drawing and writing the text for the grid
        cv2.putText(frameResized, 'A', (70, 25), cv2.FONT_HERSHEY_DUPLEX, 0.8,(255,255,255),1,cv2.LINE_AA)
        cv2.putText(frameResized, 'B', (70+160, 25), cv2.FONT_HERSHEY_DUPLEX, 0.8,(255,255,255),1,cv2.LINE_AA)
        cv2.putText(frameResized, 'C', (70+320, 25), cv2.FONT_HERSHEY_DUPLEX, 0.8,(255,255,255),1,cv2.LINE_AA)
        cv2.putText(frameResized, 'D', (70+480, 25), cv2.FONT_HERSHEY_DUPLEX, 0.8,(255,255,255),1,cv2.LINE_AA)
        cv2.putText(frameResized, 'E', (70+640, 25), cv2.FONT_HERSHEY_DUPLEX, 0.8,(255,255,255),1,cv2.LINE_AA)
        cv2.putText(frameResized, 'F', (70+800, 25), cv2.FONT_HERSHEY_DUPLEX, 0.8,(255,255,255),1,cv2.LINE_AA)
        cv2.putText(frameResized, 'G', (70+960, 25), cv2.FONT_HERSHEY_DUPLEX, 0.8,(255,255,255),1,cv2.LINE_AA)
        cv2.putText(frameResized, 'H', (70+1120, 25), cv2.FONT_HERSHEY_DUPLEX, 0.8,(255,255,255),1,cv2.LINE_AA)
        cv2.putText(frameResized, '1', (15, 80), cv2.FONT_HERSHEY_DUPLEX, 0.8,(255,255,255),1,cv2.LINE_AA)
        cv2.putText(frameResized, '2', (15, 80+144), cv2.FONT_HERSHEY_DUPLEX, 0.8,(255,255,255),1,cv2.LINE_AA)
        cv2.putText(frameResized, '3', (15, 80+288), cv2.FONT_HERSHEY_DUPLEX, 0.8,(255,255,255),1,cv2.LINE_AA)
        cv2.putText(frameResized, '4', (15, 80+432), cv2.FONT_HERSHEY_DUPLEX, 0.8,(255,255,255),1,cv2.LINE_AA)
        cv2.putText(frameResized, '5', (15, 80+576), cv2.FONT_HERSHEY_DUPLEX, 0.8,(255,255,255),1,cv2.LINE_AA)

        if status == 1:
            frame_to_write = cv2.resize(frameResized, (960, 540))
            frame_to_write = cv2.cvtColor(frame_to_write, cv2.COLOR_BGR2GRAY)
            cv2.imwrite(f"{folderpath}/frames/{frame_number}_frame.png", frame_to_write)
        
        cv2.putText(frameResized, 'Press "esc" to exit and "R" to set a new reference frame', (400, 700), cv2.FONT_HERSHEY_DUPLEX, 0.5,(200,200,200),1,cv2.LINE_AA)

        #Resize windows for stack
        gray_resized = cv2.resize(gray, (864, 486))
        thresh_resized = cv2.resize(thresh_delta, (864, 486))
        delta_resized = cv2.resize(delta_frame, (864, 486))
        black_resized = cv2.resize(black, (864, 486))

        #Stack the Windows that provide extra detail
        stack1 = numpy.hstack([gray_resized, thresh_resized])
        stack2 = numpy.hstack([delta_resized, black_resized])
        stack = numpy.vstack([stack1, stack2])
        cv2.rectangle(stack, (0,0), (864,486),(255,255,255), 1)
        cv2.rectangle(stack, (0,486), (864,971),(255,255,255), 1)
        cv2.rectangle(stack, (864,0), (1727,486),(255,255,255), 1)
        cv2.putText(stack, 'Difference', (350, 520), cv2.FONT_HERSHEY_DUPLEX, 0.8,(255,255,255),1,cv2.LINE_AA)
        cv2.putText(stack, 'Capturing', (350, 30), cv2.FONT_HERSHEY_DUPLEX, 0.8,(255,255,255),1,cv2.LINE_AA)
        cv2.putText(stack, 'Threshold', (1250, 30), cv2.FONT_HERSHEY_DUPLEX, 0.8,(255,255,255),1,cv2.LINE_AA)

        #Opening windows for visualization
        cv2.imshow('AMOS - Analysation',frameResized)
        #cv2.imshow('Additional details - AMOS', stack)

        key = cv2.waitKey(1)

        if key == 27:
            if status == 1:
                times.append(time)

                current_meteor_end = base_time + timedelta(seconds = current_seconds)

                cell = sheet.cell(row_number,5)
                cell.value = str(beginning_video_time + timedelta(seconds = current_seconds))[11:]

                cell = sheet.cell(row_number,7)
                cell.value = str(current_meteor_end)[11:]

                cell = sheet.cell(row_number,8)
                cell.value = str(current_meteor_end - current_meteor_start)
            break

        if key == ord('r'):
            ref_frame = gray
            print("Frame reset because of key pressed.")

        frame_number+=1

    video.stop()
    
    #Close all Windows
    cv2.destroyWindow('AMOS - Analysation')
    cv2.destroyWindow('Additional details - AMOS')

    try:
        wb.save(f'{folderpath}/Results.xlsx')
    except PermissionError:
        Window.spreadsheet_opened_error = QMessageBox()
        Window.spreadsheet_opened_error.setWindowTitle("Spreadsheet still opened!")
        Window.spreadsheet_opened_error.setText('You have your spreadsheet still opened! Close it and press "Save spreadsheet" again, otherwise ALL DATA WILL BE LOST!')
        Window.spreadsheet_opened_error.setIcon(QMessageBox.Warning)
        Window.spreadsheet_opened_error.setStandardButtons(QMessageBox.Close)

        x = Window.spreadsheet_opened_error.exec_()

    print("Sucessfully processed video!")
    
def generate_diagram():
    #Preparing the Pandas Dataframe
    df=pandas.DataFrame(columns=["Start","End"])

    #Calculating values and plotting the Chart
    for i in range(0,len(times),2):
        print('iteration')
        df=df.append({"Start":times[i],"End":times[i+1],"Position":position_names[i],"MeteorID":MeteorID_List[i]},ignore_index=True)

    df.to_csv(f'{folderpath}/Results.csv')

    #try:
    df["Start_string"]=df["Start"].dt.strftime("%Y-%m-%d %H:%M:%S.%f")
    df["End_string"]=df["End"].dt.strftime("%Y-%m-%d %H:%M:%S.%f")

    cds=ColumnDataSource(df)

    p=figure(x_axis_type='datetime',height=100, width=500, title='Meteore', toolbar_location="below")
    p.axis.visible = False
    p.yaxis.minor_tick_line_color=None
    p.ygrid[0].ticker.desired_num_ticks=1

    hover=HoverTool(tooltips=[("Meteor-ID", "@MeteorID"),("Position", "@Position"),
                              ("Start", "@Start_string"),("End", "@End_string")])
    p.add_tools(hover)

    q=p.quad(left="Start", right="End", bottom=0, top=1, color="black", source=cds)

    output_file("Meteordiagramm.html")
    show(p)
    #except AttributeError:
        #print("No meteors found yet!")

def save_spreadsheet(Window):
    wb.save(f'{Window.folderpath}/Results.xlsx')
    
def apply_defaults(Window):
    with open("files/defaults.txt", "r") as defaults_file:
        defaults = json.loads(defaults_file.read())
    
    if defaults[0] == [] or defaults[1] == [] or defaults[2] == "None":
        set_defaults_now = QMessageBox.question(Window, "No defaults yet!", "You have not set any defaults yet. Do you want to set them now?", QMessageBox.Yes | QMessageBox.No, QMessageBox.Yes)
        if set_defaults_now == QMessageBox.Yes:
            set_defaults(Window)
    else:
        Window.videopath_list, Window.xmlpath_list, Window.folderpath = defaults
        Window.setup_video_selection(Window.videopath_list)
        Window.setup_xml_selection(Window.xmlpath_list)
        Window.setup_folder_selection(Window.folderpath)

def set_defaults(Window):
    select_video_info = QMessageBox(icon=QMessageBox.Information, text="In the following dialog, select your default video(s).")
    select_video_info.setWindowTitle("Info")
    x = select_video_info.exec_()

    Window.default_videopath_list = QFileDialog.getOpenFileNames(parent=Window, filter="MP4 Files (*.mp4)")
    Window.default_videopath_list = Window.default_videopath_list[0]
    if Window.default_videopath_list != []: #If the user didn't cancel the selection
        select_xml_info = QMessageBox(icon=QMessageBox.Information, text="In the following dialog, select your default XML(s).")
        select_xml_info.setWindowTitle("Info")
        x = select_xml_info.exec_()

        Window.default_xmlpath_list = QFileDialog.getOpenFileNames(parent=Window, filter="XML Files (*.xml)")
        Window.default_xmlpath_list = Window.default_xmlpath_list[0]
        if Window.default_xmlpath_list != []: #If the user didn't cancel the selection
            select_folder_info = QMessageBox(icon=QMessageBox.Information, text="In the following dialog, select your default folder to store results in.")
            select_folder_info.setWindowTitle("Info")
            x = select_folder_info.exec_()

            Window.default_folderpath = QFileDialog.getExistingDirectory(parent=Window)
            if Window.default_folderpath != "": #If the user didn't cancel the selection
                default_paths = [Window.default_videopath_list, Window.default_xmlpath_list, Window.default_folderpath]
                with open("files/defaults.txt", "w") as defaults_file:
                    defaults_file.write(json.dumps(default_paths))

                set_defaults_success_message = QMessageBox(icon=QMessageBox.Information, text="Defaults set succesfully!")
                set_defaults_success_message.setWindowTitle("Info")
                x = set_defaults_success_message.exec_()

def delete_defaults(Window):
    delete_continue = QMessageBox.question(Window, "Do you want to delete?", "Are you sure that you want to delete the defaults?", QMessageBox.Yes | QMessageBox.No, QMessageBox.No)

    if delete_continue == QMessageBox.Yes:
        none_content = [[], [], "None"]
        with open("files/defaults.txt", "w") as defaults_file:
            defaults_file.write(json.dumps(none_content))

        set_defaults_success_message = QMessageBox(icon=QMessageBox.Information, text="Defaults deleted succesfully!")
        set_defaults_success_message.setWindowTitle("Info")
        x = set_defaults_success_message.exec_()

def get_thumbnail(path):
    vid = FileVideoStream(path).start()
    thumbnail = vid.read()
    vid.stop()
    thumbnail = cv2.resize(thumbnail, (240, 135))
    return thumbnail